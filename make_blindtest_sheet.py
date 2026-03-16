"""
make_blindtest_sheet.py
Build BLINDTEST.xlsx from the latest OOS backtest CSV.
Upload to Google Drive → it auto-converts to Google Sheets.

Run:
    /Library/Frameworks/Python.framework/Versions/3.14/bin/python3 make_blindtest_sheet.py
"""

import os
import csv
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(__file__)
CSV_FILE   = os.path.join(BASE_DIR, "backtest results",
                          "backtest_2024-01-01_2026-01-01_TP10pct_SL1.5_MH15d.csv")
OUT_FILE   = os.path.join(BASE_DIR, "backtest results", "BLINDTEST.xlsx")

# ── Colour palette ─────────────────────────────────────────────────────────────
C_HEADER_BG  = "1A1A2E"   # dark navy
C_HEADER_FG  = "FFFFFF"
C_SECTION_BG = "16213E"   # slightly lighter navy
C_SECTION_FG = "E0E0E0"
C_WIN_BG     = "D6F4D6"   # light green
C_LOSS_BG    = "FAD7D7"   # light red
C_ACCENT     = "0F3460"   # mid navy for alternating rows
C_ALT_ROW    = "F2F6FF"   # light blue-white
C_GOLD       = "FFD700"   # gold for summary metrics
C_WHITE      = "FFFFFF"
C_DARK_TEXT  = "1A1A2E"

def hdr_font(bold=True, size=10, color=C_HEADER_FG):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def body_font(bold=False, size=10, color=C_DARK_TEXT):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def center():
    return Alignment(horizontal="center", vertical="center")

def auto_width(ws, min_width=8, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

# ── Parse CSV sections ─────────────────────────────────────────────────────────
with open(CSV_FILE, newline="") as f:
    raw = list(csv.reader(f))

def find_section(rows, label):
    for i, row in enumerate(rows):
        if row and row[0].strip() == label:
            return i
    return None

# Config block (rows 0-8 approx)
config = {}
for row in raw[:11]:
    if len(row) >= 2 and row[0] and row[1]:
        config[row[0].strip()] = row[1].strip()

# Overall performance
perf = {}
perf_start = find_section(raw, "OVERALL PERFORMANCE") + 1
for row in raw[perf_start:perf_start + 8]:
    if len(row) >= 2 and row[0] and row[1]:
        perf[row[0].strip()] = row[1].strip()

# Exit reason breakdown
exit_start = find_section(raw, "EXIT REASON BREAKDOWN")
exit_header = raw[exit_start + 1]
exit_data = []
for row in raw[exit_start + 2:]:
    if not row or not row[0]:
        break
    exit_data.append(row)

# Per-ticker
ticker_start = find_section(raw, "PER-TICKER SUMMARY")
ticker_header = raw[ticker_start + 1]
ticker_data = []
for row in raw[ticker_start + 2:]:
    if not row or not row[0]:
        break
    ticker_data.append(row)

# Confidence tier
conf_start = find_section(raw, "CONFIDENCE TIER BREAKDOWN")
conf_header = raw[conf_start + 1]
conf_data = []
for row in raw[conf_start + 2:]:
    if not row or not row[0]:
        break
    conf_data.append(row)

# Condition analysis
cond_start = find_section(raw, "CONDITION WIN RATE ANALYSIS")
cond_header = raw[cond_start + 1]
cond_data = []
for row in raw[cond_start + 2:]:
    if not row or not row[0]:
        break
    cond_data.append(row)

# Trade log
trade_start = find_section(raw, "FULL TRADE LOG")
trade_header = raw[trade_start + 1]
trade_data = []
for row in raw[trade_start + 2:]:
    if row and row[0]:
        trade_data.append(row)

print(f"  Config keys     : {len(config)}")
print(f"  Perf keys       : {len(perf)}")
print(f"  Exit rows       : {len(exit_data)}")
print(f"  Ticker rows     : {len(ticker_data)}")
print(f"  Condition rows  : {len(cond_data)}")
print(f"  Trade rows      : {len(trade_data)}")

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_view.showGridLines = False

row = 1

# Title block
ws1.row_dimensions[row].height = 30
title_cell = ws1.cell(row, 1, "BLINDTEST — OOS Blind Test  2024–2026")
title_cell.font = Font(name="Calibri", bold=True, size=16, color=C_GOLD)
title_cell.fill = fill(C_HEADER_BG)
title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws1.merge_cells(f"A{row}:G{row}")
row += 1

subtitle = ws1.cell(row, 1, "S&P 500 Universe · TP 10% · SL 1.5×ATR · Max Hold 15d · Commission 0.1%")
subtitle.font = Font(name="Calibri", size=10, color="AAAACC")
subtitle.fill = fill(C_HEADER_BG)
subtitle.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws1.merge_cells(f"A{row}:G{row}")
row += 2

# ── Overall Performance ────────────────────────────────────────────────────────
def section_header(ws, r, text, ncols=7):
    c = ws.cell(r, 1, text)
    c.font = Font(name="Calibri", bold=True, size=11, color=C_SECTION_FG)
    c.fill = fill(C_SECTION_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
    ws.row_dimensions[r].height = 20

def kv_row(ws, r, key, value, is_positive=None):
    k = ws.cell(r, 2, key)
    k.font = Font(name="Calibri", bold=True, size=10, color=C_DARK_TEXT)
    k.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    v = ws.cell(r, 4, value)
    v.font = Font(name="Calibri", bold=True, size=10, color=C_DARK_TEXT)
    v.alignment = Alignment(horizontal="left", vertical="center")
    if is_positive is True:
        v.fill = fill(C_WIN_BG)
    elif is_positive is False:
        v.fill = fill(C_LOSS_BG)
    ws.merge_cells(f"B{r}:C{r}")
    ws.merge_cells(f"D{r}:G{r}")
    ws.row_dimensions[r].height = 18

section_header(ws1, row, "OVERALL PERFORMANCE")
row += 1

perf_items = [
    ("Total Trades",    perf.get("Total Trades",    ""),  None),
    ("Winning Trades",  perf.get("Winning Trades",  ""),  True),
    ("Losing Trades",   perf.get("Losing Trades",   ""),  False),
    ("Win Rate",        perf.get("Win Rate",        ""),  None),
    ("Avg Win",         perf.get("Avg Win",         ""),  True),
    ("Avg Loss",        perf.get("Avg Loss",        ""),  False),
    ("Profit Factor",   perf.get("Profit Factor",   ""),  None),
    ("Total P&L",       perf.get("Total P&L",       ""),  None),
]
for key, val, pos in perf_items:
    kv_row(ws1, row, key, val, pos)
    row += 1
row += 1

# ── Exit Reason Breakdown ──────────────────────────────────────────────────────
section_header(ws1, row, "EXIT REASON BREAKDOWN")
row += 1

# Header row
for ci, col_name in enumerate(exit_header, 1):
    c = ws1.cell(row, ci, col_name)
    c.font = hdr_font(size=9)
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    c.border = thin_border()
row += 1

exit_colors = {
    "Stop-Loss":      C_LOSS_BG,
    "Take-Profit":    C_WIN_BG,
    "Max Hold (15d)": C_ALT_ROW,
    "End of Period":  "FFFACD",
}
for er in exit_data:
    color = exit_colors.get(er[0], C_WHITE)
    for ci, val in enumerate(er, 1):
        c = ws1.cell(row, ci, val)
        c.font = body_font()
        c.fill = fill(color)
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                vertical="center", indent=1 if ci == 1 else 0)
        c.border = thin_border()
    row += 1
row += 1

# ── Confidence Tier ────────────────────────────────────────────────────────────
section_header(ws1, row, "CONFIDENCE TIER BREAKDOWN")
row += 1
for ci, col_name in enumerate(conf_header, 1):
    c = ws1.cell(row, ci, col_name)
    c.font = hdr_font(size=9)
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    c.border = thin_border()
row += 1
for cr in conf_data:
    for ci, val in enumerate(cr, 1):
        c = ws1.cell(row, ci, val)
        c.font = body_font()
        c.fill = fill(C_ALT_ROW)
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                vertical="center", indent=1 if ci == 1 else 0)
        c.border = thin_border()
    row += 1
row += 1

# ── Condition Analysis ─────────────────────────────────────────────────────────
section_header(ws1, row, "CONDITION WIN RATE ANALYSIS")
row += 1
for ci, col_name in enumerate(cond_header, 1):
    c = ws1.cell(row, ci, col_name)
    c.font = hdr_font(size=9)
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    c.border = thin_border()
row += 1
for i, cr in enumerate(cond_data):
    bg = C_ALT_ROW if i % 2 == 0 else C_WHITE
    for ci, val in enumerate(cr, 1):
        c = ws1.cell(row, ci, val)
        c.font = body_font()
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                vertical="center", indent=1 if ci == 1 else 0)
        c.border = thin_border()
    row += 1

# Column widths
ws1.column_dimensions["A"].width = 2
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 14
ws1.column_dimensions["D"].width = 18
for col in "EFGH":
    ws1.column_dimensions[col].width = 14

# Freeze below title rows
ws1.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — PER-TICKER BREAKDOWN
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Per-Ticker")
ws2.sheet_view.showGridLines = False

ws2.row_dimensions[1].height = 28
t = ws2.cell(1, 1, "BLINDTEST — Per-Ticker P&L Breakdown")
t.font = Font(name="Calibri", bold=True, size=14, color=C_GOLD)
t.fill = fill(C_HEADER_BG)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.merge_cells(f"A1:{get_column_letter(len(ticker_header))}1")

# Header
for ci, col_name in enumerate(ticker_header, 1):
    c = ws2.cell(2, ci, col_name)
    c.font = hdr_font()
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    c.border = thin_border()
ws2.freeze_panes = "A3"

for i, tr in enumerate(ticker_data):
    r = i + 3
    # Determine row colour from P&L column (index 4)
    try:
        pnl_str = tr[4].replace("$", "").replace(",", "").replace("+", "").strip()
        pnl_val = float(pnl_str)
        bg = C_WIN_BG if pnl_val > 0 else (C_LOSS_BG if pnl_val < 0 else C_WHITE)
    except Exception:
        bg = C_WHITE

    for ci, val in enumerate(tr, 1):
        c = ws2.cell(r, ci, val)
        c.font = body_font()
        c.fill = fill(bg)
        c.alignment = Alignment(
            horizontal="center" if ci > 1 else "left",
            vertical="center",
            indent=1 if ci == 1 else 0
        )
        c.border = thin_border()
    ws2.row_dimensions[r].height = 16

auto_width(ws2, min_width=8, max_width=18)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — FULL TRADE LOG
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Trade Log")
ws3.sheet_view.showGridLines = False

ws3.row_dimensions[1].height = 28
t3 = ws3.cell(1, 1, "BLINDTEST — Full Trade Log (1,930 Trades)")
t3.font = Font(name="Calibri", bold=True, size=14, color=C_GOLD)
t3.fill = fill(C_HEADER_BG)
t3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws3.merge_cells(f"A1:{get_column_letter(len(trade_header))}1")

# Header
for ci, col_name in enumerate(trade_header, 1):
    c = ws3.cell(2, ci, col_name)
    c.font = hdr_font()
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    c.border = thin_border()
ws3.freeze_panes = "A3"

WIN_EXITS  = {"Take-Profit", "Max Hold (15d)"}  # may still be loss on max-hold
LOSS_EXITS = {"Stop-Loss"}

for i, tr in enumerate(trade_data):
    r = i + 3
    result_col = 4  # "Result" column (WIN / LOSS), 0-indexed = index 4
    try:
        result = tr[4].strip().upper()
        bg = C_WIN_BG if result == "WIN" else (C_LOSS_BG if result == "LOSS" else C_ALT_ROW)
    except Exception:
        bg = C_ALT_ROW

    for ci, val in enumerate(tr, 1):
        c = ws3.cell(r, ci, val)
        c.font = body_font(size=9)
        c.fill = fill(bg)
        c.alignment = Alignment(
            horizontal="center" if ci not in (2, 16) else "left",
            vertical="center",
            indent=1 if ci in (2, 16) else 0,
            wrap_text=(ci == 16)
        )
        c.border = thin_border()
    ws3.row_dimensions[r].height = 16

# Column widths for trade log
col_widths = {
    1: 5,   # #
    2: 7,   # Ticker
    3: 6,   # Signal
    4: 5,   # Stars
    5: 6,   # Result
    6: 12,  # Entry Date
    7: 10,  # Entry Price
    8: 10,  # Stop Loss
    9: 10,  # Target
    10: 12, # Exit Date
    11: 10, # Exit Price
    12: 8,  # Days Held
    13: 16, # Exit Reason
    14: 8,  # P&L %
    15: 10, # P&L $
    16: 40, # Strategies
}
for ci, w in col_widths.items():
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — WORST TICKERS (negative P&L, sorted by loss)
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Worst Performers")
ws4.sheet_view.showGridLines = False

ws4.row_dimensions[1].height = 28
t4 = ws4.cell(1, 1, "BLINDTEST — Worst Performing Tickers (Negative P&L)")
t4.font = Font(name="Calibri", bold=True, size=14, color=C_GOLD)
t4.fill = fill(C_HEADER_BG)
t4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws4.merge_cells(f"A1:{get_column_letter(len(ticker_header))}1")

for ci, col_name in enumerate(ticker_header, 1):
    c = ws4.cell(2, ci, col_name)
    c.font = hdr_font()
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()
    c.border = thin_border()
ws4.freeze_panes = "A3"

def pnl_sort_key(row):
    try:
        return float(row[4].replace("$", "").replace(",", "").replace("+", "").strip())
    except Exception:
        return 0

worst = sorted(
    [tr for tr in ticker_data if tr[4].startswith("$-") or tr[4].startswith("-")],
    key=pnl_sort_key
)

for i, tr in enumerate(worst):
    r = i + 3
    # gradient: deeper red for bigger losses
    for ci, val in enumerate(tr, 1):
        c = ws4.cell(r, ci, val)
        c.font = body_font()
        c.fill = fill(C_LOSS_BG)
        c.alignment = Alignment(
            horizontal="center" if ci > 1 else "left",
            vertical="center",
            indent=1 if ci == 1 else 0
        )
        c.border = thin_border()
    ws4.row_dimensions[r].height = 16

auto_width(ws4, min_width=8, max_width=18)

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(OUT_FILE)
print(f"\n✅  Saved: {OUT_FILE}")
print(f"    Sheets: Summary | Per-Ticker | Trade Log | Worst Performers")
print(f"\nTo open in Google Sheets:")
print(f"  1. Go to drive.google.com → New → File upload")
print(f"  2. Upload BLINDTEST.xlsx")
print(f"  3. Right-click → Open with → Google Sheets")
print(f"  4. File → Rename → 'BLINDTEST'")
